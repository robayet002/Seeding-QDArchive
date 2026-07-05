"""
Export the classification result table to an XLSX spreadsheet
(Classification_instruction).

Columns:
    repository_id | project_type | project_title |
    primary_class | secondary_class | no_project_files

no_project_files = total number of files in the project.

Run:  python export_xlsx.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import REPORT_XLSX_PATH
from database import get_connection

HEADERS = [
    "repository_id",
    "project_type",
    "project_title",
    "primary_class",
    "secondary_class",
    "no_project_files",
]

QUERY = """
    SELECT
        p.repository_id                          AS repository_id,
        p.type                                   AS project_type,
        p.title                                  AS project_title,
        pc.primary_class_name                    AS primary_class,
        pc.secondary_class_name                  AS secondary_class,
        (SELECT COUNT(*) FROM files f
          WHERE f.project_id = p.id)             AS no_project_files
    FROM projects p
    LEFT JOIN project_classification pc ON pc.project_id = p.id
    ORDER BY p.repository_id, p.type, p.title
"""


def export_classification_table(output_path=REPORT_XLSX_PATH):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(QUERY)
    rows = cursor.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Classification"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="4472C4")
    body_font = Font(name="Arial")

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append(["" if v is None else v for v in row])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font

    # Reasonable column widths
    widths = [14, 16, 60, 55, 55, 16]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

    wb.save(output_path)
    print(f"XLSX exported to: {output_path}  ({len(rows)} rows)")
    return output_path


if __name__ == "__main__":
    export_classification_table()
